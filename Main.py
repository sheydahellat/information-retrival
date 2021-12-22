import math
import pickle
from pathlib import Path
from parsivar import FindStems
import stopWords
import openpyxl
import itertools
from itertools import islice

import matplotlib.pyplot as plt
from collections import OrderedDict

punc = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''
xlsx_file = Path('../IR1_7k_news.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
score = [0] * 8000


def normalize_query(words):
    my_stemmer = FindStems()
    stemFree = []
    for word in stopWords.stopWords:
        for token in words:
            if token == word:
                words.remove(token)
    for word in words:
        stemFree.append(my_stemmer.convert_to_stem(word))
    return stemFree


def index_elimination(normalized_query):
    unique_list= []
    same = []

    for q in normalized_query:
        if q not in unique_list:
            unique_list.append(q)

    for term in unique_list:
        count_q_query = unique_list.count(term)
        wqt = 1+math.log(count_q_query, 10)
        idf = len(pos_index[term][1])
        if(term not in pos_index):
            continue
        for d in pos_index[term][1]:
            count = pos_index[term][1][d][0]
            wd = (1 + math.log(count, 10)) * math.log(7562 / idf, 10)
            score[d] += (wd*wqt)

        li = []
        for i in range(len(score)):
            li.append([score[i], i])
        li.sort()
        sort_index = []

        for x in li:
            sort_index.append(x[1])
        same = Reverse(sort_index)[0:10]

    return same

def search_championList(normalized_query):
    unique_list= []
    same = []

    for q in normalized_query:
        if q not in unique_list:
            unique_list.append(q)

    for term in unique_list:
        # print(pos_index[term][1])
        count_q_query = normalized_query.count(term)
        wqt = 1+math.log(count_q_query, 10)
        if(term not in pos_index):
            continue

        for d in pos_index[term][1]:
            # print(d)
            wd = pos_index[term][1][d]
            score[d] += (wd*wqt)


        li = []
        for i in range(len(score)):
            li.append([score[i], i])
        # print(li)
        li.sort()
        sort_index = []
        # print(li)

        for x in li:
            sort_index.append(x[1])
        same = Reverse(sort_index)[0:10]

    return same

def Reverse(lst):
    new_lst = lst[::-1]
    return new_lst



#reading from file
posindex_file = open("championList.pkl", "rb")
pos_index = pickle.load(posindex_file)

#get query & normalaizing


print("enter your query:")
query = input()
for p in punc:
    query = query.replace(p, "")
sep_query=query.split()
normalized_query = normalize_query(sep_query)
# print(normalized_query)
same = search_championList(normalized_query)
for doc_id in same:
    print(wb_obj.active.cell(doc_id + 1, 2).value)
    print(wb_obj.active.cell(doc_id + 1, 3).value)
